# -*- coding: utf-8 -*-

# To use the script, put the 2 files in the same directory as the script. First file is the new transactions and second is the dashboard
# ex: python extractor.py transactions.xlsx DepensesSalaire.xlsx

import pandas as pd
import calendar
from openpyxl import load_workbook
import os, sys

#%% Import transactions

transacFile = os.path.basename(sys.argv[1])
transac = pd.read_excel(transacFile)
# transac = pd.read_excel(r'./transactions.xlsx')

DepensesSalaireFile = os.path.basename(sys.argv[2])

#%% Extract expenses and incomes from transactions
incomes = transac.loc[transac["Montant"] >= 0 ]

expenses = transac.loc[transac["Montant"] < 0 ]

#%% Save expenses in the Excel Dashboard

# Take only interresting data for the dashboard
filtered_expenses = pd.DataFrame(expenses, columns = ["Date comptable", "Montant", "Compte de la contrepartie", "Nom de la contrepartie", "Communication"])

# Add month and year from the "Date Compatable" in the the DataFrame
dates = pd.DataFrame(filtered_expenses, columns=["Date comptable"])
dates['Year'] =  dates["Date comptable"].dt.year
dates['Month'] =  dates["Date comptable"].dt.month
dates['Month'] = dates['Month'].apply(lambda x: calendar.month_abbr[x])
filtered_expenses.insert(1, "Year", dates["Year"])
filtered_expenses.insert(2, "Month", dates["Month"])

# Remove time from the date
# filtered_expenses["Date comptable"] = pd.to_datetime(dates['Date comptable']).dt.date

# Sort the dates in the reverse to match with the excel dashboard 
filtered_expenses = filtered_expenses.sort_values(by=["Date comptable"], ascending=True)
filtered_expenses = filtered_expenses.reset_index(drop=True)
#%% Save incomes in the Excel Dashboard

# Take only interresting data for the dashboard
filtered_incomes = pd.DataFrame(incomes, columns = ["Date comptable", "Montant", "Compte de la contrepartie", "Nom de la contrepartie", "Communication"])

# Add month and year from the "Date Compatable" in the the DataFrame
dates = pd.DataFrame(filtered_incomes, columns=["Date comptable"])
dates['Year'] =  dates["Date comptable"].dt.year
dates['Month'] =  dates["Date comptable"].dt.month
dates['Month'] = dates['Month'].apply(lambda x: calendar.month_abbr[x])
filtered_incomes.insert(1, "Year", dates["Year"])
filtered_incomes.insert(2, "Month", dates["Month"])

# Remove time from the date
# filtered_incomes["Date comptable"] = pd.to_datetime(dates['Date comptable']).dt.date

# Sort the dates in the reverse to match with the excel dashboard 
filtered_incomes = filtered_incomes.sort_values(by=["Date comptable"], ascending=True)
filtered_incomes = filtered_incomes.reset_index(drop=True)

#%% Save expenses in the last row of the dashboard document (.xlsx)

# Load file & open the correct sheet
path = DepensesSalaireFile
# path = "DepensesSalaireFile.xlsx"
wb = load_workbook(path)
ws = wb["Depenses"]

# Recover the max_row when open the file (last expense in the dashboard)
last_row = ws.max_row

# Take the last date of the dashboard to only add new expenses
last_date = ws.cell(row=ws.max_row,column=1).value

# Recover the last expense of the dashboard to compare with new transactions and only add new expense
last_expense = []

for col in ws.iter_cols(min_row=ws.max_row, max_row=ws.max_row, max_col=7):
  for cell in col:    
    # Fill the None and NAN value of last expense with 0 to allow comparison just after with each expense    
    if cell.value == None:
        last_expense.append(0)
    else:    
        last_expense.append(cell.value)

# Check if the last expense is in the new transaction. If not the case, add directly the new expense
# If it's the case, récupérer seulement les expense qui sont après. Supprimer tout les précédents filtered expense déjà ajoutés.
for index, each_expense in filtered_expenses.iterrows():
    # Fill the None and NAN value of each expense with 0 to allow comparison with last expense
    each_expense = each_expense.fillna(0)
    
    if each_expense.tolist() == last_expense:
        index_to_remove = index
        filtered_expenses = filtered_expenses.truncate(before=index_to_remove+1)

# Iterate on each rows pf the dataframe (new expenses)
step = 1

for index, each_expense in filtered_expenses.iterrows():
    # Iterate on each column cell to add new expense from the dataframe only if new expense
    for element in range(1,8):
        ws.cell(row=last_row+step, column=element).value = each_expense[element-1]
    step += 1

# Save the xlsx
wb.save(path)

#%% Save incomes in the last row of the dashboard document (.xlsx) BON

# Load file and open the correct sheet
path = DepensesSalaireFile
# path = "DepensesSalaireFile.xlsx"
wb = load_workbook(path)
ws2 = wb["Incomes"]

# Recover the max_row when open the file (last income in the dashboard)
last_row = ws2.max_row

# Take the last date of the dashboard to only add new incomes
last_date = ws2.cell(row=ws2.max_row,column=1).value


# Recover the last income of the dashboard to compare with new transactions and only add new incomes
last_income = []

for col in ws2.iter_cols(min_row=ws2.max_row, max_row=ws2.max_row, max_col=7):
  for cell in col:
    # Fill the None and NAN value of last income with 0 to allow comparison just after with each income    
    if cell.value == None:
        last_income.append(0)
    else:    
        last_income.append(cell.value)

# Check if the last income is in the new transaction. If not the case, add directly the new incomes
# If it's the case, récupérer seulement les incomes qui sont après. Supprimer tout les précédents filtered incomes déjà ajoutés.
for index, each_income in filtered_incomes.iterrows():
    # Fill the None and NAN value of each income with 0 to allow comparison with last income
    each_income = each_income.fillna(0)
    
    if each_income.tolist() == last_income:
        index_to_remove = index
        filtered_incomes = filtered_incomes.truncate(before=index_to_remove+1)

# Iterate on each rows of the dataframe (new incomes)
step = 1

for index, each_income in filtered_incomes.iterrows():
    # Iterate on each column cell to add new incomes from the dataframe only if new income
    for element in range(1,8):
        ws2.cell(row=last_row+step, column=element).value = each_income[element-1]
    step += 1

# Save the xlsx
wb.save(path)